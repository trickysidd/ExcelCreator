import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, StringVar, font
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from PIL import Image as PilImage
import tempfile
import shutil
import threading
from datetime import date
from pathlib import Path


class ExcelImageProcessor(tk.Tk):
    def __init__(self):
        super().__init__()
        print(self)
        self.title("Excel Image Processor")
        self.geometry("600x190")
        self.resizable(False, False)
        
        self.create_widgets()
        
        self.option_add('*Dialog.msg.font', 'Calibri 10')
        self.option_add('*Dialog.msg.wrapLength',400)
        self.error_log = []
        self.cancel_event = threading.Event()

    def validate_integer(self,new_text):
        self.new_text = new_text
        if self.new_text.isdigit() and len(self.new_text) <=3:
            return True
        if self.new_text == "":
            return True
        return False

    def show_help(self):
        self.help_message = (
            "this is help"
            )
        messagebox.showinfo("Help",self.help_message)
        

        
    def ask_directory(self,entry_var,status_var,which):
        """Open a directory selection dialog and update the text box variable."""
        self.directory = filedialog.askdirectory()
        self.entry_var = entry_var
        self.status_var = status_var
        if self.directory:
            self.entry_var.set(self.directory)
            if which == "source":
                self.status_var.set('Workdone Folder Selected')
                self.button_dir2.config(state='normal')
            elif which == "dest":
                self.status_var.set('Save Folder Selected')
                self.button_start.config(state='normal')
        else:
            print("No directory Selected")
            self.status_var.set('No Folder Selected')

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(expand=True, fill='both')
        vcmd = (self.register(self.validate_integer),'%P')
        self.status_message = StringVar()
        self.status_message.set("Waiting to Start")
        

        # --- Row 0: First directory selector ---##
        self.frame_row0 = ttk.Frame(main_frame)
        self.frame_row0.pack(fill='x', padx=10, pady=5)
        self.button_dir1 = ttk.Button(self.frame_row0, text="Folder")
        self.entry_dir1_var = StringVar()
        self.entry_dir1_var.set("Please Select the Parent Folder containing all Workdone Folders")
        self.entry_dir1 = ttk.Entry(self.frame_row0, textvariable=self.entry_dir1_var, state='readonly')
        # Command for the button uses a lambda to pass the StringVar
        self.button_dir1.config(command=lambda: self.ask_directory(self.entry_dir1_var,self.status_message, 'source'))
        self.button_dir1.pack(side='left', padx=(0, 5))
        self.entry_dir1.pack(side='left', fill='x', expand=True)
        #---- Row0 Finish-----##


        
        # --- Row 1: Width and Height inputs ---
        self.frame_row1 = ttk.Frame(main_frame)
        self.frame_row1.pack(fill='x', padx=10, pady=5)
        self.label_width = ttk.Label(self.frame_row1, text="Image Width")
        self.entry_width = ttk.Entry(self.frame_row1,width=4,validate='key',validatecommand=vcmd)
        self.entry_width.insert(0,"300")
        self.label_height = ttk.Label(self.frame_row1, text="Image Height")
        self.entry_height = ttk.Entry(self.frame_row1,width=4,validate='key',validatecommand=vcmd)
        self.entry_height.insert(0,"200")
        self.button_help = ttk.Button(self.frame_row1, text="Help")
        self.label_width.pack(side='left')
        self.entry_width.pack(side='left', fill='x', expand=False, padx=(5, 10))
        self.label_height.pack(side='left')
        self.entry_height.pack(side='left', fill='x', expand=False, padx=(5, 0))
        self.button_help.pack(side='right')
        self.button_help.config(command=lambda: self.show_help())
        #------Row1 Finish------#
        
        # --- Row 2: Second directory selector and Start button ---
        self.frame_row2 = ttk.Frame(main_frame)
        self.frame_row2.pack(fill='x', padx=10, pady=5)
        self.button_dir2 = ttk.Button(self.frame_row2, text="Save Folder",state='disabled')
        self.entry_dir2_var = StringVar()
        self.entry_dir2 = ttk.Entry(self.frame_row2,state='readonly', textvariable=self.entry_dir2_var)
        self.button_start = ttk.Button(self.frame_row2, text="Start", command=self.on_start_button_click,state='disabled')
        self.button_cancel = ttk.Button(self.frame_row2, text="Cancel", command=self.on_cancel_button_click,state='normal')
        # Command for the button uses a lambda to pass the StringVar
        self.button_dir2.config(command=lambda: self.ask_directory(self.entry_dir2_var,self.status_message, 'dest'))
        self.button_dir2.pack(side='left', padx=(0, 5))
        self.entry_dir2.pack(side='left', fill='x', expand=True)
        self.button_start.pack(side='right', padx=(5, 0))
        #----Row2 Finish-------#
        
        # --- Row 3: Third row Status ---
        self.frame_row3 = ttk.Frame(main_frame)
        self.frame_row3.pack(fill='x', padx=10, pady=5)
        self.underlined_arial_font = font.Font(family='Arial',slant='italic',size=10,underline=True)
        self.label_status = ttk.Label(self.frame_row3,text="Status : ")
        self.label_current= ttk.Label(self.frame_row3,textvariable=self.status_message,font=self.underlined_arial_font)
        self.label_status.pack(side='left',padx=(0,5))
        self.label_current.pack(side='left',padx=(0,5))
        #-----Row3 Finish-----#

        
        self.progress_bar = ttk.Progressbar(main_frame, orient='horizontal', mode='indeterminate', length=360)

     
        
       # start_button = ttk.Button(main_frame, text="Start", command=self.on_start_button_click)
       # start_button.pack(pady=10)

       
    def on_cancel_button_click(self):
        self.cancel_event.set()
        messagebox.showerror("Error","Operation Cancelled")
        self.status_message.set("Last Operation was Cancelled by the User, Reselect all Folders")




        
    def on_start_button_click(self):
        # Disable the button and show progress bar during processing
        for widget in self.winfo_children():        
            for child in widget.winfo_children():       
                for children in child.winfo_children():                 
                    if isinstance(children, ttk.Button):
                        children.config(state='disabled')
                        
        self.button_start.pack_forget()
        self.button_cancel.pack(side='right', padx=(5, 0))
        self.button_cancel.config(state='normal')
        self.progress_bar.pack(pady=5)
        self.progress_bar.start()
        #print(self.entry_width.get())

        # Run the long-running task in a separate thread to keep the GUI responsive

        processing_thread = threading.Thread(target=self.create_excel_process,args=(self.entry_dir1_var.get(),
                            self.entry_width.get(),self.entry_height.get(),self.entry_dir2_var.get(),self.cancel_event))

        processing_thread.start()
        print("Processing started")

    def update_status(self, message):
        self.status_label.config(text=message)
        self.update_idletasks() # Force an immediate GUI update

    

    def reset_gui(self):
        self.progress_bar.stop()
        self.progress_bar.pack_forget()
        for widget in self.winfo_children():
            for child in widget.winfo_children():
                for children in child.winfo_children():
                    if isinstance(children, ttk.Button):
                        children.config(state='normal')
        self.button_cancel.pack_forget()
        self.button_start.pack(side='right', padx=(5, 0))
        self.button_start.config(state='disabled')
        
    def create_excel_process(self,source_path,img_width,img_height,save_directory,cancel_event):
        try:
            root_folder_path = source_path
##            root_folder_path = filedialog.askdirectory(title="Select the root folder with images")
            Today = date.today().strftime("%d %B %Y")
            
            if not root_folder_path:
                self.status_message.set("No folder selected.")
                #self.reset_gui()
                return

            root_folder_path = os.path.abspath(root_folder_path)
            output_excel_name = f'DryDock Engine Room Workdone for {Today}.xlsx'
            img_width = int(img_width)
            img_height = int(img_height)
            wb = Workbook()
            ws = wb.active
            


            '''Styles'''
            thick_border_side = Side(style='thick')
            thick_border = Border(left=thick_border_side, right=thick_border_side, top=thick_border_side, bottom=thick_border_side)
            thin_side = Side(style='thin')
            normal_all_borders= Border(left=thin_side,right=thin_side,top=thin_side,bottom=thin_side)
            center_alignment = Alignment(horizontal="center",vertical='center',wrapText=True)
            title_font = Font(size=24, bold=True)
            work_font = Font(size=14)
            '''Styles'''
            ws.merge_cells("A1:C1")
            ws.merge_cells("A2:C2")
            ws.merge_cells("A3:C3")
            ws.row_dimensions[2].height = 30
            ws["A2"] = f'Engine Room Workdone for {Today}'
            ws["A2"].alignment = center_alignment
            ws["A2"].font = title_font
            ws["A2"].border = thick_border
            
            
            row_num = 4
            col_num = 1
            max_size = (img_width, img_height) # Maximum size for images in the Excel file

            # Use a context manager to handle temporary directory creation and cleanup
            with tempfile.TemporaryDirectory() as tmp_dir:
                # Loop through all subfolders in the root directory
                for dirpath, dirnames, filenames in os.walk(root_folder_path):
                    # Exclude the root folder itself and process only subfolders
                    if cancel_event.is_set():
                        break
                    if dirpath == root_folder_path:
                        continue

                    # Extract the subfolder name
                    subfolder_name = os.path.basename(dirpath)
                    print(f'Now Processing {subfolder_name} Folder')
                    self.status_message.set(f'Processing {subfolder_name} Folder')
                    
                    # Write the subfolder name as a header in the first column
                    ws.merge_cells(f'A{row_num}:C{row_num}')
                    topic_cell = ws.cell(row=row_num, column=1, value=f"{subfolder_name}: Write Job details here")
                    topic_cell.alignment = Alignment(wrapText=True)
                    topic_cell.font = work_font
                    topic_cell.border = thick_border
                    row_num += 1

                    # Get image files from the current subfolder
                    image_files = [f for f in filenames if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif'))]
                    
                    # Loop through image files and insert them below the subfolder name
                    for filename in image_files:
                        if cancel_event.is_set():
                            break
                        image_path = os.path.join(dirpath, filename)
                        
                        try:
                            # Create a unique temporary path for each image
                            temp_resized_path = os.path.join(tmp_dir, f"resized_{filename}")
                            
                            # Load, resize, and save the image
                            img_pil = PilImage.open(image_path)
                            img_pil.thumbnail(max_size, PilImage.Resampling.LANCZOS)
                            img_pil.save(temp_resized_path)
                            
                            # Create an Openpyxl Image object
                            img_openpyxl = OpenpyxlImage(temp_resized_path)
                            cur_col = chr(64+col_num)
                            # Insert the image into the worksheet
                            cell_address = f"{cur_col}{row_num}"

                            '''New'''
                            offset_x = offset_y =30000
                            _from = AnchorMarker(col=col_num-1, row=row_num - 1,colOff=offset_x,rowOff=offset_y)
                            to = AnchorMarker(col=col_num,row=row_num,colOff=-offset_x,rowOff=-offset_y)
                            img_openpyxl.anchor = TwoCellAnchor(editAs="twoCell",_from=_from, to=to)
                            ws.add_image(img_openpyxl)
                            '''New Finish'''

                            
                            #ws.add_image(img_openpyxl, cell_address)
                            
                            #ws[cell_address].border = thick_border

                            # Set row height and column width
                            ws.column_dimensions[chr(64 + col_num)].width = img_openpyxl.width / 7
                            ws.row_dimensions[row_num].height = img_openpyxl.height * 0.8
                            
                            
                            col_num += 1 
                            if col_num > 3:
                                col_num = 1
                                row_num += 1

                        except Exception as e:
                            print(f"Error processing image {filename} in subfolder {subfolder_name}: {e}")
                            self.error_log.append(e)
                            
                    if cancel_event.is_set():
                        break    
                    
                    # Add an extra space after each folder's images for better readability
                    row_num += 1
                    col_num = 1

                max_row = ws.max_row
                max_col = ws.max_column
                for row_numb in range(1,max_row +1):
                    for col_numb in range(1,max_col+1):
                        cell = ws.cell(row=row_numb,column=col_numb)
                        cell.border = normal_all_borders

                if not cancel_event.is_set():
            # Save the workbook
                    try:
                        save_directory= save_directory
                        #save_directory = filedialog.askdirectory(title="Please Select the Destination")
                        wb.save(os.path.join(save_directory,output_excel_name))
                        print(f"Excel file '{output_excel_name}' created successfully.")
                        self.status_message.set(f'Workdone File saved in {self.entry_dir2_var.get()}')
                        self.reset_gui()
                    except Exception as e:
                        print(f"Error saving the Excel file: {e}")
                        self.error_log.append(e)
                    
        except Exception as e:
                    print(f"Error saving the Excel file: {e}")
        finally:
            self.reset_gui()
            #self.status_message.set(f'Workdone File saved in {self.entry_dir2_var.get()}')
            print(self.error_log)



if __name__ == "__main__":
    app = ExcelImageProcessor()
    app.mainloop()
